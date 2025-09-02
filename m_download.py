#!/usr/bin/env python3
import os, time, zipfile, shutil, sys, re
from pathlib import Path
from typing import Optional, List, Tuple

from openpyxl import load_workbook

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException, WebDriverException

# ---------------- CONFIG ----------------
EXCEL_PATH = "sharepoint_folders.xlsx"
URL_HEADER  = "FolderURL"
QID_HEADER  = "QID"
SHEET_NAME  = None  # e.g. "Sheet1" or None → first sheet

OUTPUT_DIR = Path("OUTPUT")
RUN_DOWNLOAD_DIR = Path("session_downloads")

# Point to your existing Chrome profile (macOS examples)
CHROME_USER_DATA_DIR = os.path.expanduser("~/Library/Application Support/Google/Chrome")
CHROME_PROFILE_DIR_NAME = "Default"  # or "Profile 1", etc.

PAGE_LOAD_TIMEOUT = 60
CLICK_WAIT_TIMEOUT = 40
DOWNLOAD_MAX_WAIT = 600
DOWNLOAD_STABLE_SECONDS = 3
DELETE_ZIP_AFTER_EXTRACT = True
HEADLESS = False  # set True if you want headless (not recommended for UI-heavy pages)

# ---------------------------------------


# ===== Excel helpers =====
def _norm(s: Optional[str]) -> str:
    return (s or "").strip().lower()

def load_url_qid_rows(xlsx_path: str, url_header: str, qid_header: str, sheet_name: Optional[str]) -> List[Tuple[str, str]]:
    wb = load_workbook(filename=xlsx_path, data_only=True, read_only=False)
    ws = wb[sheet_name] if sheet_name else wb.active

    header_map = {}
    for idx, cell in enumerate(ws[1], start=1):
        if cell.value is not None:
            header_map[_norm(str(cell.value))] = idx

    if _norm(url_header) not in header_map or _norm(qid_header) not in header_map:
        raise ValueError(f"Headers not found. Have: {list(header_map.keys())}")

    url_col = header_map[_norm(url_header)]
    qid_col = header_map[_norm(qid_header)]

    rows = []
    for r in range(2, ws.max_row + 1):
        url_cell = ws.cell(row=r, column=url_col)
        qid_cell = ws.cell(row=r, column=qid_col)

        qid_val = qid_cell.value
        if qid_val is None or str(qid_val).strip() == "":
            continue

        url_val = None
        if url_cell.hyperlink and getattr(url_cell.hyperlink, "target", None):
            url_val = url_cell.hyperlink.target
        elif isinstance(url_cell.value, str):
            raw = url_cell.value.strip()
            if raw.lower().startswith(("http://", "https://")):
                url_val = raw

        if url_val:
            rows.append((url_val, str(qid_val).strip()))

    wb.close()
    return rows


# ===== Driver =====
def build_driver(download_dir: Path) -> webdriver.Chrome:
    download_dir.mkdir(parents=True, exist_ok=True)
    opts = Options()
    opts.add_argument(f"--user-data-dir={CHROME_USER_DATA_DIR}")
    opts.add_argument(f"--profile-directory={CHROME_PROFILE_DIR_NAME}")
    if HEADLESS:
        opts.add_argument("--headless=new")

    prefs = {
        "download.default_directory": str(download_dir.resolve()),
        "download.prompt_for_download": False,
        "download.directory_upgrade": True,
        "safebrowsing.enabled": True,
        "plugins.always_open_pdf_externally": True,
    }
    opts.add_experimental_option("prefs", prefs)

    drv = webdriver.Chrome(options=opts)
    drv.set_page_load_timeout(PAGE_LOAD_TIMEOUT)
    return drv


# ===== Download helpers =====
def has_in_progress(download_dir: Path) -> bool:
    return any(p.suffix == ".crdownload" for p in download_dir.iterdir())

def newest_after(download_dir: Path, after_ts: float) -> Optional[Path]:
    best, best_m = None, after_ts
    for p in download_dir.iterdir():
        try:
            m = p.stat().st_mtime
            if m > best_m:
                best, best_m = p, m
        except Exception:
            pass
    return best

def wait_for_download_complete(download_dir: Path, start_ts: float, timeout: int, stable_seconds: int) -> Optional[Path]:
    deadline = time.time() + timeout
    candidate = None
    last_size = -1
    stable_since = None

    while time.time() < deadline:
        newest = newest_after(download_dir, after_ts=start_ts - 1)
        if newest:
            candidate = newest
        if has_in_progress(download_dir):
            time.sleep(0.5)
            continue
        if candidate and candidate.exists():
            size = candidate.stat().st_size
            if size != last_size:
                last_size = size
                stable_since = time.time()
            else:
                if stable_since and (time.time() - stable_since) >= stable_seconds:
                    return candidate
        time.sleep(0.5)
    return None


# ===== Post-process =====
def qid_dir(base: Path, qid: str) -> Path:
    d = base / f"QID_{qid}"
    d.mkdir(parents=True, exist_ok=True)
    return d

def extract_zip(zip_path: Path, dest: Path):
    with zipfile.ZipFile(zip_path, "r") as zf:
        zf.extractall(dest)

def move_file(src: Path, dest_dir: Path) -> Path:
    dest = dest_dir / src.name
    if dest.exists():
        stem, ext = dest.stem, dest.suffix
        i = 2
        while True:
            cand = dest_dir / f"{stem} ({i}){ext}"
            if not cand.exists():
                dest = cand
                break
            i += 1
    shutil.move(str(src), str(dest))
    return dest


# ===== UI Automation strategies =====
DOWNLOAD_XPATHS = [
    # Modern SharePoint/OneDrive
    "//button[@aria-label='Download']",
    "//button[@title='Download']",
    "//*[@data-automationid='download']",
    "//div[@role='menuitem' and contains(@aria-label,'Download')]",
    "//span[normalize-space()='Download']/ancestor::button[1]",
    "//span[contains(.,'Download')]/ancestor::button[1]",
    # Overflow / command bar item
    "//*[@role='menuitem' and .//span[normalize-space()='Download']]",
]

MORE_MENU_XPATHS = [
    # Overflow 'More' / '…' menu in command bar
    "//button[@aria-label='More']",
    "//button[@aria-label='More actions']",
    "//button[contains(@aria-label,'More')]",
    "//button[.//span[contains(.,'More')]]",
]

def js_click_first_download_like(driver) -> bool:
    """Last-resort: query all clickable elements, click the first visible one whose text ~ 'Download'."""
    js = r"""
    const matches = [];
    const walker = document.createTreeWalker(document.body, NodeFilter.SHOW_ELEMENT, null);
    while (walker.nextNode()) {
      const el = walker.currentNode;
      try {
        const style = window.getComputedStyle(el);
        if (style && style.display !== 'none' && style.visibility !== 'hidden' && el.offsetParent !== null) {
          const text = (el.innerText || el.textContent || '').trim();
          if (/^\s*download\s*$/i.test(text) || /download/i.test(text)) {
            matches.push(el);
          }
        }
      } catch(e) {}
    }
    if (matches.length) {
      matches[0].click();
      return true;
    }
    return false;
    """
    try:
        return bool(driver.execute_script(js))
    except Exception:
        return False

def try_click_many(driver, xpaths: List[str], wait_sec: int) -> bool:
    wait = WebDriverWait(driver, wait_sec)
    last_err = None
    for xp in xpaths:
        try:
            el = wait.until(EC.element_to_be_clickable((By.XPATH, xp)))
            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
            time.sleep(0.2)
            el.click()
            return True
        except Exception as e:
            last_err = e
    if last_err:
        # try JS scan as fallback
        return js_click_first_download_like(driver)
    return False

def select_all_items(driver):
    # SharePoint often respects Ctrl+A / Cmd+A to select items in list view
    try:
        body = driver.find_element(By.TAG_NAME, "body")
        if sys.platform == "darwin":
            body.send_keys(Keys.COMMAND, "a")
        else:
            body.send_keys(Keys.CONTROL, "a")
        time.sleep(0.5)
    except Exception:
        pass

def in_iframes_click_download(driver) -> bool:
    # Some viewers load inside iframes (PDF or file preview). Iterate all iframes.
    frames = driver.find_elements(By.TAG_NAME, "iframe")
    for i, f in enumerate(frames):
        try:
            driver.switch_to.frame(f)
            if try_click_many(driver, DOWNLOAD_XPATHS, 3):
                driver.switch_to.default_content()
                return True
            # Last resort inside frame
            if js_click_first_download_like(driver):
                driver.switch_to.default_content()
                return True
        except Exception:
            pass
        finally:
            try:
                driver.switch_to.default_content()
            except Exception:
                pass
    return False

def try_direct_download_url(driver, url: str) -> bool:
    # Append download=1 (avoid duplicating)
    if "download=1" not in url.lower():
        sep = "&" if ("?" in url) else "?"
        direct = url + f"{sep}download=1"
    else:
        direct = url
    try:
        driver.get(direct)
        return True
    except Exception:
        return False

def trigger_download(driver) -> bool:
    """
    Attempt multiple strategies to trigger a download on the current page.
    Returns True if we *think* a download was initiated.
    """
    # 1) Toolbar Download
    if try_click_many(driver, DOWNLOAD_XPATHS, CLICK_WAIT_TIMEOUT):
        return True
    # 2) Try opening overflow 'More' then Download
    if try_click_many(driver, MORE_MENU_XPATHS, 5):
        if try_click_many(driver, DOWNLOAD_XPATHS, 8):
            return True
    # 3) Select all then Download
    select_all_items(driver)
    if try_click_many(driver, DOWNLOAD_XPATHS, 6):
        return True
    # 4) Iframes (preview/download button inside frame)
    if in_iframes_click_download(driver):
        return True
    # 5) JS text-based click fallback
    if js_click_first_download_like(driver):
        return True
    return False


# ===== Main =====
def main():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    if RUN_DOWNLOAD_DIR.exists():
        for p in RUN_DOWNLOAD_DIR.iterdir():
            if p.is_file():
                p.unlink()
            else:
                shutil.rmtree(p)
    RUN_DOWNLOAD_DIR.mkdir(parents=True, exist_ok=True)

    rows = load_url_qid_rows(EXCEL_PATH, URL_HEADER, QID_HEADER, SHEET_NAME)
    print(f"Found {len(rows)} rows. Starting…")

    driver = build_driver(RUN_DOWNLOAD_DIR)
    try:
        ok = 0
        fails: List[str] = []

        for i, (url, qid) in enumerate(rows, 1):
            print(f"\n[{i}/{len(rows)}] QID={qid}")
            start_ts = time.time()

            # Try direct download URL first (fast path)
            direct_ok = try_direct_download_url(driver, url)
            if not direct_ok:
                # Fall back to normal navigation
                try:
                    driver.get(url)
                except Exception as e:
                    fails.append(f"QID={qid} navigation failed: {e}")
                    continue

            # Try to trigger a download via UI if direct didn't auto-start
            # (Even for direct_ok we attempt UI in case the site ignored the flag.)
            triggered = trigger_download(driver)
            if not triggered:
                # It might still be a direct download (no UI to click). Proceed to wait.
                pass

            finished = wait_for_download_complete(
                RUN_DOWNLOAD_DIR, start_ts,
                timeout=DOWNLOAD_MAX_WAIT,
                stable_seconds=DOWNLOAD_STABLE_SECONDS
            )
            if not finished:
                fails.append(f"QID={qid} download did not complete (timeout).")
                continue

            target = qid_dir(OUTPUT_DIR, qid)
            if finished.suffix.lower() == ".zip":
                try:
                    extract_zip(finished, target)
                    if DELETE_ZIP_AFTER_EXTRACT:
                        finished.unlink(missing_ok=True)
                    print(f"  -> Extracted to {target}")
                    ok += 1
                except Exception as e:
                    fails.append(f"QID={qid} unzip failed: {e}")
            else:
                try:
                    dest = move_file(finished, target)
                    print(f"  -> Saved file to {dest}")
                    ok += 1
                except Exception as e:
                    fails.append(f"QID={qid} move failed: {e}")

        print(f"\nDone. Success: {ok}/{len(rows)}")
        if fails:
            print("\nFailures:")
            for f in fails:
                print(" -", f)

    finally:
        try:
            driver.quit()
        except Exception:
            pass


if __name__ == "__main__":
    main()
