#!/usr/bin/env python3
import os
import time
import subprocess
from pathlib import Path
from datetime import datetime
from typing import Optional, List, Tuple
import urllib.parse as urlparse
import zipfile
import shutil
from openpyxl import load_workbook

# --------------- CONFIG ---------------
EXCEL_PATH = "sharepoint_folders.xlsx"    # your Excel file
COLUMN_URL_HEADER = "FolderURL"           # header name for the column that has hyperlinks
COLUMN_QID_HEADER = "QID"                 # header name for QID
OUTPUT_BASE = Path("output")              # where extracted content goes
DOWNLOADS_DIR = Path.home() / "Downloads"
PER_FOLDER_TIMEOUT = 600                  # seconds to wait for each folder ZIP
STABLE_SECONDS = 4                        # how long a file size must be stable
OPEN_IN_BACKGROUND = False                # don't bring Safari to front if True
CLOSE_TAB_AFTER = True                    # try to close tab after triggering download
RENAME_FILES_REPLACE_SPACES = True        # rename extracted files/dirs: " " -> "_"
# --------------------------------------


def force_download_param(u: str) -> str:
    """Append ?download=1 (or &download=1) unless already present."""
    try:
        parsed = urlparse.urlparse(u)
        q = urlparse.parse_qsl(parsed.query, keep_blank_values=True)
        if not any(k.lower() == "download" for k, _ in q):
            q.append(("download", "1"))
        new_q = urlparse.urlencode(q)
        return urlparse.urlunparse(parsed._replace(query=new_q))
    except Exception:
        return u


def applescript_open_in_safari(url: str, activate: bool = True) -> None:
    """Open URL in a new Safari tab using AppleScript."""
    script = f'''
    set theURL to "{url.replace('"', '%22')}"
    tell application "Safari"
        if (count of windows) = 0 then
            make new document with properties {{URL:theURL}}
        else
            tell window 1 to set current tab to (make new tab with properties {{URL:theURL}})
        end if
        {"activate" if activate else ""}
    end tell
    '''
    subprocess.run(["osascript", "-e", script], check=True)


def applescript_close_active_tab() -> None:
    """Close the current tab in the front Safari window."""
    script = '''
    tell application "Safari"
        if (count of windows) > 0 then
            tell window 1
                if (count of tabs) > 0 then
                    close current tab
                end if
            end tell
        end if
    end tell
    '''
    subprocess.run(["osascript", "-e", script], check=True)


def newest_path(dirpath: Path, after_ts: float) -> Optional[Path]:
    newest = None
    newest_mtime = after_ts
    for p in dirpath.iterdir():
        try:
            m = p.stat().st_mtime
            if m > newest_mtime:
                newest_mtime = m
                newest = p
        except Exception:
            pass
    return newest


def looks_in_progress(p: Path) -> bool:
    """Safari in-progress downloads appear as .download packages (directories)."""
    if p.suffix.lower() == ".download":
        return True
    if not p.is_file():
        return True
    return False


def wait_until_stable(file_path: Path, stable_seconds: int, timeout_s: int) -> bool:
    start = time.time()
    last_size = -1
    last_change = time.time()
    while time.time() - start < timeout_s:
        if not file_path.exists():
            time.sleep(0.5)
            continue
        try:
            size = file_path.stat().st_size
        except Exception:
            time.sleep(0.5)
            continue

        if size != last_size:
            last_size = size
            last_change = time.time()
        else:
            if time.time() - last_change >= stable_seconds:
                return True
        time.sleep(0.5)
    return False


def unzip_to_target(zip_path: Path, target_dir: Path):
    target_dir.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(zip_path, 'r') as zf:
        zf.extractall(target_dir)

    if RENAME_FILES_REPLACE_SPACES:
        for root, dirs, files in os.walk(target_dir, topdown=False):
            for name in files:
                if " " in name:
                    src = Path(root) / name
                    dst = Path(root) / name.replace(" ", "_")
                    if not dst.exists():
                        src.rename(dst)
            for name in dirs:
                if " " in name:
                    src = Path(root) / name
                    dst = Path(root) / name.replace(" ", "_")
                    if not dst.exists():
                        src.rename(dst)


def process_folder(url: str, qid_value) -> tuple[bool, str]:
    """Open folder link in Safari, trigger ZIP download, extract to QID folder."""
    qid_str = str(qid_value).strip()
    if not qid_str:
        return False, f"Empty QID for URL: {url}"
    target_dir = OUTPUT_BASE / f"QID_{qid_str}"

    start_ts = time.time()
    folder_url = force_download_param(url)

    try:
        applescript_open_in_safari(folder_url, activate=(not OPEN_IN_BACKGROUND))
    except subprocess.CalledProcessError as e:
        return False, f"Safari open failed: {e}"

    deadline = start_ts + PER_FOLDER_TIMEOUT

    while time.time() < deadline:
        newest = newest_path(DOWNLOADS_DIR, after_ts=start_ts - 1)
        if newest is None:
            time.sleep(0.5)
            continue

        if looks_in_progress(newest):
            time.sleep(0.5)
            continue

        # If we get a single file instead of a ZIP, still store it under the QID dir.
        if newest.suffix.lower() != ".zip":
            try:
                target_dir.mkdir(parents=True, exist_ok=True)
                dest = target_dir / newest.name
                wait_until_stable(newest, STABLE_SECONDS, timeout_s=int(max(1, deadline - time.time())))
                shutil.move(str(newest), str(dest))
                if RENAME_FILES_REPLACE_SPACES and " " in dest.name:
                    dest.rename(dest.with_name(dest.name.replace(" ", "_")))
                if CLOSE_TAB_AFTER:
                    try: applescript_close_active_tab()
                    except Exception: pass
                return True, f"Saved single file -> {dest}"
            except Exception as e:
                if CLOSE_TAB_AFTER:
                    try: applescript_close_active_tab()
                    except Exception: pass
                return False, f"Non-zip download handling failed: {e}"

        # ZIP path: wait to stabilize then extract
        if wait_until_stable(newest, STABLE_SECONDS, timeout_s=int(max(1, deadline - time.time()))):
            try:
                unzip_to_target(newest, target_dir)
                newest.unlink(missing_ok=True)  # remove zip after extraction
                if CLOSE_TAB_AFTER:
                    try: applescript_close_active_tab()
                    except Exception: pass
                return True, f"Extracted to {target_dir}"
            except Exception as e:
                if CLOSE_TAB_AFTER:
                    try: applescript_close_active_tab()
                    except Exception: pass
                return False, f"Unzip failed: {e}"
        else:
            break

    if CLOSE_TAB_AFTER:
        try: applescript_close_active_tab()
        except Exception: pass
    return False, f"Timeout waiting for folder ZIP from: {url}"


def _normalize_header(s: Optional[str]) -> str:
    return (s or "").strip().lower()


def _looks_like_url(s: str) -> bool:
    s = s.strip().lower()
    return s.startswith("http://") or s.startswith("https://")


def load_rows_from_excel(xlsx_path: str,
                         url_header: str,
                         qid_header: str) -> List[Tuple[str, str]]:
    """
    Reads (url, qid) pairs from an .xlsx file, preferring the cell's hyperlink target over
    the displayed text. Returns list of (url, qid).
    """
    wb = load_workbook(filename=xlsx_path, data_only=True, read_only=False)
    ws = wb.active  # or choose by name: wb["Sheet1"]

    # Find header row (assume first non-empty row is header)
    header_row_idx = None
    for r in range(1, min(ws.max_row, 20) + 1):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        if any(v is not None and str(v).strip() for v in row_vals):
            header_row_idx = r
            break
    if header_row_idx is None:
        return []

    # Map headers to column indices
    header_map = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row_idx, column=c).value
        header_map[_normalize_header(str(v) if v is not None else "")] = c

    url_col = header_map.get(_normalize_header(url_header))
    qid_col = header_map.get(_normalize_header(qid_header))
    if not url_col or not qid_col:
        raise ValueError(
            f"Could not find columns '{url_header}' and/or '{qid_header}'. "
            f"Found headers: {list(header_map.keys())}"
        )

    rows: List[Tuple[str, str]] = []
    for r in range(header_row_idx + 1, ws.max_row + 1):
        url_cell = ws.cell(row=r, column=url_col)
        qid_cell = ws.cell(row=r, column=qid_col)

        # Extract QID
        qid_val = qid_cell.value
        if qid_val is None or str(qid_val).strip() == "":
            continue

        # Prefer the cell hyperlink target, if present
        url_val = None
        if url_cell.hyperlink and getattr(url_cell.hyperlink, "target", None):
            url_val = url_cell.hyperlink.target
        else:
            # fallback to cell text if it looks like a URL
            raw = url_cell.value
            if isinstance(raw, str) and _looks_like_url(raw):
                url_val = raw

        if url_val:
            rows.append((url_val.strip(), str(qid_val).strip()))

    wb.close()
    return rows


def main():
    OUTPUT_BASE.mkdir(parents=True, exist_ok=True)
    if not Path(EXCEL_PATH).exists():
        print(f"Excel not found: {EXCEL_PATH}")
        return

    try:
        rows = load_rows_from_excel(EXCEL_PATH, COLUMN_URL_HEADER, COLUMN_QID_HEADER)
    except Exception as e:
        print(f"Failed to read hyperlinks from Excel: {e}")
        return

    if not rows:
        print("No (URL, QID) pairs found. "
              "Make sure your sheet has headers and hyperlinks in the URL column.")
        return

    print(f"Found {len(rows)} rows with folder links and QIDs. Starting…")

    ok = 0
    fails = []
    for i, (url, qid) in enumerate(rows, 1):
        t0 = datetime.now().strftime("%H:%M:%S")
        print(f"[{i}/{len(rows)} {t0}] QID={qid} :: {url}")
        success, msg = process_folder(url, qid)
        print("  ->", msg)
        ok += int(success)
        if not success:
            fails.append(f"QID={qid} :: {url} :: {msg}")

    print(f"\nDone. Success: {ok}/{len(rows)}")
    if fails:
        print("\nFailures (check manually or adjust):")
        for f in fails:
            print(" -", f)


if __name__ == "__main__":
    main()
